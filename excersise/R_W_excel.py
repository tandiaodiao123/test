#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2020/7/2 11:04
# @Author  : datan
# @Site    : 
# @File    : R_W_excel.py

from openpyxl import load_workbook
from test_request import http_request
#读取数据
def read_data(file_name,sheet_name):
    wb = load_workbook(file_name)
    sheet = wb[sheet_name]
    all_case=[]#所有测试用例读取
    for i in range(1,sheet.max_row):
        case=[]
        for j in range(1,sheet.max_column-1):
            case.append(sheet.cell(row=i+1,column=j).value)
        all_case.append(case)
    return all_case

# 写入数据
Token=''#定义全局变量
def write_data(file,file_name,sheet_name,c1,c2):
    global Token
    for test_data in  file :
        data_response=http_request(test_data[3],eval(test_data[4]),test_data[2],token=Token)#接收返回测试数据
        token=data_response.json()
        if 'login' in test_data[3]:
            test_data_2=file[0]
            token_response = http_request(test_data_2[3], eval(test_data_2[4]), test_data_2[2], token=Token)
            token=token_response.json()['data']['token_info']['token']
            Token='Bearer '+token
        # 写入测试返回数据
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]
        value=sheet.cell(row=test_data[0]+1,column=c1).value=str(data_response.json())
        # 写入结果
        result={'code':data_response.json()['code'],'msg':data_response.json()['msg']}
        expected=eval(test_data[5])
        if expected==result:
            value_2 = sheet.cell(row=test_data[0] + 1, column=c2).value = 'PASS'
        else:
            value_2 = sheet.cell(row=test_data[0] + 1, column=c2).value = 'Fail'
        wb.save(file_name)





